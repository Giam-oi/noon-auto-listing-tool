from __future__ import annotations

from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill

from .models import ProductDraft


ADJUSTMENT_COLUMNS = ["sku", "title_en", "category_key", "UAE_stock", "KSA_stock", "notes"]


def write_stock_adjustment_workbook(path: Path, drafts: list[ProductDraft]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "Stock Adjustments"
    ws.append(ADJUSTMENT_COLUMNS)
    for cell in ws[1]:
        cell.fill = PatternFill("solid", fgColor="1F4E78")
        cell.font = Font(color="FFFFFF", bold=True)
    for draft in drafts:
        ws.append(
            [
                draft.sku,
                draft.listing.title_en,
                draft.category.category_key,
                draft.stock.get("UAE", 1000),
                draft.stock.get("KSA", 1000),
                "",
            ]
        )
    ws.freeze_panes = "A2"
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 60
    ws.column_dimensions["C"].width = 22
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 14
    ws.column_dimensions["F"].width = 30
    wb.save(path)


def read_stock_adjustments(path: Path) -> dict[str, dict[str, int]]:
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    headers = [str(cell.value or "").strip() for cell in ws[1]]
    index = {name: idx for idx, name in enumerate(headers)}
    result: dict[str, dict[str, int]] = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        sku = str(row[index.get("sku", 0)] or "").strip()
        if not sku:
            continue
        result[sku] = {}
        for market in ["UAE", "KSA"]:
            col = index.get(f"{market}_stock")
            if col is None or col >= len(row):
                continue
            value = row[col]
            try:
                result[sku][market] = int(value)
            except (TypeError, ValueError):
                continue
    wb.close()
    return result


def apply_stock_to_workbook(path: Path, adjustments: dict[str, dict[str, int]]) -> int:
    wb = load_workbook(path)
    ws = wb[wb.sheetnames[0]]
    headers = [str(ws.cell(1, col).value or "").strip() for col in range(1, ws.max_column + 1)]
    try:
        sku_col = headers.index("sku") + 1
        market_col = headers.index("marketplace") + 1
        stock_col = headers.index("stock") + 1
    except ValueError:
        wb.close()
        return 0
    changed = 0
    for row in range(2, ws.max_row + 1):
        sku = str(ws.cell(row, sku_col).value or "").strip()
        market = str(ws.cell(row, market_col).value or "").strip()
        value = adjustments.get(sku, {}).get(market)
        if value is not None:
            ws.cell(row, stock_col).value = value
            changed += 1
    wb.save(path)
    wb.close()
    return changed


def apply_stock_to_run(run_dir: Path, adjustment_path: Path) -> dict[str, Any]:
    adjustments = read_stock_adjustments(adjustment_path)
    exports = run_dir / "exports"
    changed: dict[str, int] = {}
    for name in ["review_workbook.xlsx", "noon_bulk_UAE.xlsx", "noon_bulk_KSA.xlsx"]:
        path = exports / name
        if path.exists():
            changed[name] = apply_stock_to_workbook(path, adjustments)
    return {"run_dir": str(run_dir), "adjustment_path": str(adjustment_path), "changed": changed}
