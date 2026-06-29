from __future__ import annotations

import json
import re
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any, Callable
from urllib.parse import urlparse

from openpyxl import load_workbook


URL_PATTERN = re.compile(r"https?://[^\s\"'<>]+", re.IGNORECASE)


@dataclass
class BatchItem:
    source: str
    status: str = "pending"
    run_dir: str = ""
    error: str = ""
    summary: dict[str, Any] = field(default_factory=dict)


@dataclass
class BatchResult:
    total: int
    succeeded: int = 0
    needs_review: int = 0
    failed: int = 0
    items: list[BatchItem] = field(default_factory=list)

    def to_dict(self) -> dict[str, Any]:
        return {
            "total": self.total,
            "succeeded": self.succeeded,
            "needs_review": self.needs_review,
            "failed": self.failed,
            "items": [
                {
                    "source": item.source,
                    "status": item.status,
                    "run_dir": item.run_dir,
                    "error": item.error,
                    "summary": item.summary,
                }
                for item in self.items
            ],
        }


class BatchRunner:
    def __init__(self, pipeline: Any):
        self.pipeline = pipeline

    def _pipeline_for_item(self) -> Any:
        if callable(self.pipeline) and not hasattr(self.pipeline, "build_from_1688_url"):
            return self.pipeline()
        return self.pipeline

    def _emit(self, item: BatchItem, progress_callback: Callable[[BatchItem], None] | None) -> None:
        if progress_callback:
            progress_callback(
                BatchItem(
                    source=item.source,
                    status=item.status,
                    run_dir=item.run_dir,
                    error=item.error,
                    summary=dict(item.summary),
                )
            )

    def _validation_issue_summary(self, run_dir: str) -> str:
        if not run_dir:
            return ""
        path = Path(run_dir) / "validation_report.json"
        if not path.exists():
            return ""
        try:
            data = json.loads(path.read_text(encoding="utf-8"))
        except (OSError, json.JSONDecodeError):
            return ""
        issues: list[str] = []
        for draft in data.get("drafts", []):
            for issue in draft.get("issues", []):
                if issue.get("severity") == "error":
                    code = str(issue.get("code") or "").strip()
                    message = str(issue.get("message") or "").strip()
                    if code and message:
                        issues.append(f"{code}: {message}")
                    elif code:
                        issues.append(code)
        return "; ".join(issues[:5])

    def run_urls(
        self,
        urls: list[str],
        progress_callback: Callable[[BatchItem], None] | None = None,
    ) -> BatchResult:
        result = BatchResult(total=len(urls))
        for url in urls:
            item = BatchItem(source=url, status="running", error="Collecting product data")
            self._emit(item, progress_callback)
            try:
                summary = self._pipeline_for_item().build_from_1688_url(url)
                item.summary = dict(summary)
                item.run_dir = str(summary.get("run_dir", ""))
                if int(summary.get("submit_ready") or 0) > 0:
                    item.status = "succeeded"
                    item.error = ""
                    result.succeeded += 1
                else:
                    item.status = "needs_review"
                    item.error = self._validation_issue_summary(item.run_dir)
                    result.needs_review += 1
            except Exception as exc:
                item.status = "failed"
                item.error = str(exc)
                result.failed += 1
            result.items.append(item)
            self._emit(item, progress_callback)
        return result


def _canonical_1688_url(value: str) -> str:
    parsed = urlparse(value.strip().rstrip(".,;"))
    if parsed.scheme not in {"http", "https"}:
        return ""
    if not parsed.netloc.lower().endswith("1688.com"):
        return ""
    match = re.search(r"/offer/(\d+)\.html", parsed.path)
    if not match:
        return ""
    return f"https://detail.1688.com/offer/{match.group(1)}.html"


def parse_1688_urls(text: str) -> list[str]:
    seen: set[str] = set()
    urls: list[str] = []
    for match in URL_PATTERN.finditer(text or ""):
        canonical = _canonical_1688_url(match.group(0))
        if canonical and canonical not in seen:
            seen.add(canonical)
            urls.append(canonical)
    return urls


def read_urls_from_excel(path: str | Path) -> list[str]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        found: list[str] = []
        for worksheet in workbook.worksheets:
            for row in worksheet.iter_rows(values_only=True):
                for value in row:
                    if isinstance(value, str):
                        found.extend(parse_1688_urls(value))
        return parse_1688_urls("\n".join(found))
    finally:
        workbook.close()
