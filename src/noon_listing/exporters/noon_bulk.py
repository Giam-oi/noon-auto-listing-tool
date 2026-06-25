from __future__ import annotations

import csv
import json
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from ..models import ProductDraft
from ..stock_adjustments import write_stock_adjustment_workbook


GENERIC_COLUMNS = [
    "sku",
    "parent_sku",
    "marketplace",
    "category_path",
    "category_key",
    "brand",
    "model_number",
    "model_name",
    "title_en",
    "title_ar",
    "bullets_en",
    "bullets_ar",
    "description_en",
    "description_ar",
    "search_keywords",
    "price",
    "currency",
    "stock",
    "main_image",
    "image_2",
    "image_3",
    "image_4",
    "image_5",
    "source_url",
    "attributes_json",
    "submit_ready",
    "submit_score",
    "validation_issues",
]


def _draft_market_row(draft: ProductDraft, market: str) -> dict[str, Any]:
    price = draft.prices[market]
    attrs = draft.listing.attributes
    images = list(draft.images)
    issues = [f"{i.severity}:{i.code}:{i.field}:{i.message}" for i in draft.validation_issues]
    return {
        "sku": draft.sku,
        "parent_sku": draft.parent_sku,
        "marketplace": market,
        "category_path": draft.category.noon_path,
        "category_key": draft.category.category_key,
        "brand": attrs.get("brand", "Generic"),
        "model_number": attrs.get("model_number", ""),
        "model_name": attrs.get("model_name", ""),
        "title_en": draft.listing.title_en,
        "title_ar": draft.listing.title_ar,
        "bullets_en": "\n".join(draft.listing.bullets_en),
        "bullets_ar": "\n".join(draft.listing.bullets_ar),
        "description_en": draft.listing.description_en,
        "description_ar": draft.listing.description_ar,
        "search_keywords": ", ".join(draft.listing.search_keywords),
        "price": price.list_price,
        "currency": price.currency,
        "stock": draft.stock.get(market, 0),
        "main_image": images[0] if len(images) > 0 else "",
        "image_2": images[1] if len(images) > 1 else "",
        "image_3": images[2] if len(images) > 2 else "",
        "image_4": images[3] if len(images) > 3 else "",
        "image_5": images[4] if len(images) > 4 else "",
        "source_url": draft.source_product.source_url,
        "attributes_json": json.dumps(attrs, ensure_ascii=False),
        "submit_ready": draft.submit_ready,
        "submit_score": draft.submit_score,
        "validation_issues": "\n".join(issues),
    }


def _style_sheet(ws) -> None:
    fill = PatternFill("solid", fgColor="1F4E78")
    font = Font(color="FFFFFF", bold=True)
    for cell in ws[1]:
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.freeze_panes = "A2"
    widths = {
        "A": 20,
        "B": 20,
        "C": 12,
        "D": 42,
        "I": 52,
        "J": 42,
        "K": 48,
        "M": 60,
        "S": 55,
        "AB": 70,
    }
    for idx, column in enumerate(ws.iter_cols(1, ws.max_column), start=1):
        letter = get_column_letter(idx)
        if letter in widths:
            ws.column_dimensions[letter].width = widths[letter]
        else:
            ws.column_dimensions[letter].width = min(28, max(12, len(str(ws.cell(1, idx).value)) + 2))
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)


def write_generic_workbook(path: Path, rows: list[dict[str, Any]], sheet_name: str = "Noon Bulk") -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    ws.append(GENERIC_COLUMNS)
    for row in rows:
        ws.append([row.get(col, "") for col in GENERIC_COLUMNS])
    _style_sheet(ws)
    wb.save(path)


def write_generic_csv(path: Path, rows: list[dict[str, Any]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with open(path, "w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=GENERIC_COLUMNS)
        writer.writeheader()
        writer.writerows(rows)


def write_review_workbook(path: Path, drafts: list[ProductDraft], markets: list[str]) -> None:
    rows = []
    for draft in drafts:
        for market in markets:
            rows.append(_draft_market_row(draft, market))
    write_generic_workbook(path, rows, "Review")


def fill_downloaded_template(template_path: Path, output_path: Path, rows: list[dict[str, Any]]) -> bool:
    if not template_path.exists():
        return False
    wb = load_workbook(template_path)
    ws = wb[wb.sheetnames[0]]
    header_row = None
    headers: list[str] = []
    for row_idx in range(1, min(12, ws.max_row) + 1):
        values = [str(ws.cell(row_idx, col).value or "").strip() for col in range(1, ws.max_column + 1)]
        hits = sum(1 for v in values if _map_template_header(v))
        if hits >= 3:
            header_row = row_idx
            headers = values
            break
    if header_row is None:
        wb.close()
        return False
    start_row = header_row + 1
    for offset, row in enumerate(rows):
        excel_row = start_row + offset
        for col_idx, header in enumerate(headers, start=1):
            mapped = _map_template_header(header)
            if mapped:
                ws.cell(excel_row, col_idx).value = row.get(mapped, "")
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    wb.close()
    return True


def _map_template_header(header: str) -> str:
    h = header.lower().strip().replace(" ", "_")
    mapping = {
        "seller_sku": "sku",
        "sku": "sku",
        "partner_sku": "sku",
        "parent_sku": "parent_sku",
        "title": "title_en",
        "title_en": "title_en",
        "english_title": "title_en",
        "arabic_title": "title_ar",
        "title_ar": "title_ar",
        "brand": "brand",
        "model_number": "model_number",
        "model_name": "model_name",
        "description": "description_en",
        "description_en": "description_en",
        "price": "price",
        "stock": "stock",
        "quantity": "stock",
        "image": "main_image",
        "main_image": "main_image",
        "image_url_1": "main_image",
        "image_url_2": "image_2",
        "image_url_3": "image_3",
        "image_url_4": "image_4",
        "image_url_5": "image_5",
    }
    return mapping.get(h, "")


def export_all(drafts: list[ProductDraft], out_dir: Path, cfg: dict[str, Any]) -> dict[str, str]:
    out_dir.mkdir(parents=True, exist_ok=True)
    markets = list(cfg.get("marketplaces", {}).keys())
    outputs: dict[str, str] = {}
    review = out_dir / "review_workbook.xlsx"
    write_review_workbook(review, drafts, markets)
    outputs["review_workbook"] = str(review)
    stock_adjustments = out_dir / "stock_adjustments.xlsx"
    write_stock_adjustment_workbook(stock_adjustments, drafts)
    outputs["stock_adjustments"] = str(stock_adjustments)
    template_dir = Path(cfg.get("noon", {}).get("template_dir", "templates"))
    if not template_dir.is_absolute():
        template_dir = Path(__file__).resolve().parents[2] / template_dir
    for market in markets:
        rows = [_draft_market_row(draft, market) for draft in drafts]
        xlsx = out_dir / f"noon_bulk_{market}.xlsx"
        csv_path = out_dir / f"noon_bulk_{market}.csv"
        template = template_dir / f"noon_{market.lower()}_electronics_template.xlsx"
        templated = out_dir / f"noon_bulk_{market}_template.xlsx"
        used_template = fill_downloaded_template(template, templated, rows)
        if used_template:
            outputs[f"noon_bulk_{market}_template"] = str(templated)
        write_generic_workbook(xlsx, rows, f"Noon {market}")
        write_generic_csv(csv_path, rows)
        outputs[f"noon_bulk_{market}"] = str(xlsx)
        outputs[f"noon_bulk_{market}_csv"] = str(csv_path)
    return outputs
