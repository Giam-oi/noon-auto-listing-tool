from __future__ import annotations

from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from .models import SourceProduct


FIELD_ALIASES = {
    "sku": ["sku", "product sku", "product_sku", "产品sku", "产品SKU", "SKU"],
    "title_cn": ["title", "name", "product", "product name", "产品", "产品名称", "中文名"],
    "price_cny": ["price", "cost", "采购单价", "单价", "成本", "成本价"],
    "stock": ["stock", "库存", "数量", "采购数量"],
    "source_url": ["url", "link", "source_url", "1688", "1688链接", "链接"],
    "description_cn": ["description", "desc", "描述", "卖点", "详情"],
    "image": ["image", "image path", "图片", "图片路径", "产品图"],
}


def _norm(value: Any) -> str:
    return str(value or "").strip().lower().replace(" ", "").replace("_", "")


def _header_map(headers: list[Any]) -> dict[str, int]:
    normalized = [_norm(v) for v in headers]
    result: dict[str, int] = {}
    for field, aliases in FIELD_ALIASES.items():
        alias_norms = [_norm(v) for v in aliases]
        for idx, value in enumerate(normalized):
            if value in alias_norms:
                result[field] = idx
                break
    return result


def _find_header_row(rows: list[list[Any]]) -> tuple[int, dict[str, int]]:
    best_idx = 0
    best_map: dict[str, int] = {}
    for idx, row in enumerate(rows[:8]):
        mapping = _header_map(row)
        if len(mapping) > len(best_map):
            best_idx = idx
            best_map = mapping
    return best_idx, best_map


def read_products_from_excel(path: Path, sheet: str | int | None = None, max_rows: int | None = None) -> list[SourceProduct]:
    wb = load_workbook(path, read_only=True, data_only=True)
    if sheet is None:
        ws = wb[wb.sheetnames[0]]
    elif isinstance(sheet, int):
        ws = wb[wb.sheetnames[sheet]]
    else:
        ws = wb[sheet]
    rows = [list(row) for row in ws.iter_rows(values_only=True)]
    wb.close()
    if not rows:
        return []
    header_idx, mapping = _find_header_row(rows)
    products: list[SourceProduct] = []
    for row in rows[header_idx + 1 :]:
        if max_rows and len(products) >= max_rows:
            break
        def get(field: str) -> Any:
            idx = mapping.get(field)
            if idx is None or idx >= len(row):
                return None
            return row[idx]

        title = str(get("title_cn") or "").strip()
        source_url = str(get("source_url") or "").strip()
        if not title and not source_url:
            continue
        price_raw = get("price_cny")
        try:
            price_cny = float(price_raw) if price_raw not in (None, "") else None
        except (TypeError, ValueError):
            price_cny = None
        product = SourceProduct(
            source="excel",
            source_url=source_url,
            title_cn=title,
            description_cn=str(get("description_cn") or "").strip(),
            price_cny=price_cny,
        )
        sku = str(get("sku") or "").strip()
        if sku:
            product.attributes["source_sku"] = sku
        image = str(get("image") or "").strip()
        if image and Path(image).exists():
            product.local_images.append(str(Path(image).resolve()))
        products.append(product)
    return products
